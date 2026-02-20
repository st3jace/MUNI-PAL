"""
Excel chunking service.

Per spec: Excel chunks are sheet-based.
"""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from munipal.services.chunking.base import BaseChunker, ChunkData

logger = logging.getLogger(__name__)


class ExcelChunker(BaseChunker):
    """
    Extracts chunks from Excel workbooks.

    Each sheet becomes one chunk with:
    - sheet_name preserved
    - text_content as formatted table
    - content_hash computed for deduplication
    """

    def __init__(self, max_rows: int = 1000, max_cols: int = 50):
        """
        Initialize Excel chunker.

        Args:
            max_rows: Maximum rows to extract per sheet
            max_cols: Maximum columns to extract per sheet
        """
        self.max_rows = max_rows
        self.max_cols = max_cols

    def chunk(self, file_path: Path) -> list[ChunkData]:
        """
        Extract sheet-based chunks from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of ChunkData, one per sheet
        """
        chunks: list[ChunkData] = []

        try:
            # Load workbook in read-only mode for efficiency
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            logger.info(f"Processing Excel with {len(sheet_names)} sheets: {file_path.name}")

            for seq_num, sheet_name in enumerate(sheet_names):
                sheet = wb[sheet_name]

                # Extract sheet content as text
                text_content = self._extract_sheet_text(sheet)
                text_content = self.clean_text(text_content)

                # Generate content hash
                hash_content = text_content or f"sheet_{sheet_name}_empty"
                content_hash = self.compute_hash(hash_content)

                # Get sheet dimensions
                try:
                    row_count = sheet.max_row or 0
                    col_count = sheet.max_column or 0
                except Exception:
                    row_count = 0
                    col_count = 0

                chunk = ChunkData(
                    chunk_type="sheet",
                    sequence_number=seq_num,
                    text_content=text_content,
                    content_hash=content_hash,
                    sheet_name=sheet_name,
                    metadata={
                        "total_sheets": len(sheet_names),
                        "row_count": row_count,
                        "col_count": col_count,
                        "has_content": text_content is not None,
                    },
                )
                chunks.append(chunk)

            wb.close()
            logger.info(f"Extracted {len(chunks)} chunks from Excel")

        except Exception as e:
            logger.error(f"Failed to process Excel {file_path}: {e}")
            raise ValueError(f"Excel processing failed: {e}")

        return chunks

    def _extract_sheet_text(self, sheet) -> str | None:
        """
        Extract text content from a sheet.

        Formats the sheet as a structured table with row numbers for readability.
        The format is designed to be easily interpreted by AI for financial data extraction.
        """
        rows_data = []

        try:
            row_num = 0
            for row in sheet.iter_rows(max_row=self.max_rows, max_col=self.max_cols):
                row_num += 1
                row_values = []
                for cell in row:
                    value = cell.value
                    if value is not None:
                        # Convert to string and clean
                        str_value = str(value).strip()
                        # Truncate very long cell values
                        if len(str_value) > 500:
                            str_value = str_value[:500] + "..."
                        row_values.append(str_value)
                    else:
                        row_values.append("")

                # Only add row if it has some content
                if any(v for v in row_values):
                    # Add row number prefix for easier reference
                    row_text = " | ".join(row_values)
                    rows_data.append(f"Row {row_num}: {row_text}")

        except Exception as e:
            logger.warning(f"Error extracting sheet content: {e}")
            return None

        if not rows_data:
            return None

        # Add header explaining the format
        header = "[Excel spreadsheet data - each row is pipe-delimited (|), first column typically contains labels]"
        return header + "\n" + "\n".join(rows_data)


class CSVChunker(BaseChunker):
    """
    Extracts chunks from CSV files.

    The entire CSV becomes one chunk (since CSVs don't have sheets).
    """

    def __init__(self, max_rows: int = 5000):
        self.max_rows = max_rows

    def chunk(self, file_path: Path) -> list[ChunkData]:
        """Extract a single chunk from a CSV file."""
        import csv

        chunks: list[ChunkData] = []

        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows_data = []
                row_count = 0

                for row in reader:
                    if row_count >= self.max_rows:
                        break

                    # Join row values with pipe delimiter
                    row_text = " | ".join(str(cell).strip() for cell in row)
                    if row_text.strip():
                        rows_data.append(row_text)

                    row_count += 1

            text_content = self.clean_text("\n".join(rows_data)) if rows_data else None
            content_hash = self.compute_hash(text_content or "empty_csv")

            chunk = ChunkData(
                chunk_type="sheet",  # Treat CSV as single sheet
                sequence_number=0,
                text_content=text_content,
                content_hash=content_hash,
                sheet_name="data",
                metadata={
                    "row_count": row_count,
                    "has_content": text_content is not None,
                },
            )
            chunks.append(chunk)

            logger.info(f"Extracted chunk from CSV with {row_count} rows")

        except Exception as e:
            logger.error(f"Failed to process CSV {file_path}: {e}")
            raise ValueError(f"CSV processing failed: {e}")

        return chunks
